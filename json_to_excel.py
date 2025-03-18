import sys
import os
import json
import pandas as pd
import datetime
import itertools
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def get_excel_filename(json_folder):
    """
    Determines the Excel file name based on company_info.json.
    If a valid company name is available, uses it.
    Otherwise, names the file as company_data_[Date]_[Time].xlsx.
    """
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    default_filename = f"company_data_{timestamp}.xlsx"

    try:
        company_info_path = os.path.join(json_folder, "company_info.json")
        if os.path.exists(company_info_path):
            with open(company_info_path, "r") as f:
                company_data = json.load(f)

            # Check for valid company name
            if (isinstance(company_data, dict) and 
                "Name" in company_data and 
                company_data["Name"] and 
                isinstance(company_data["Name"], str) and 
                company_data["Name"].strip()):
                
                # Sanitize the company name
                sanitized_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in company_data["Name"].strip())
                if sanitized_name:  # Make sure we still have a name after sanitizing
                    excel_filename = f"{sanitized_name}_{timestamp}.xlsx"
                    print(f"📁 Excel file will be saved as: {excel_filename}")
                    return excel_filename

        print(f"📁 Using default filename: {default_filename}")
        return default_filename

    except Exception as e:
        print(f"⚠️ Error generating filename: {e}. Using default filename.")
        return default_filename

def normalize_json_data(data):
    """
    Ensures all lists in the JSON data have the same length by filling missing values with empty strings.
    """
    if isinstance(data, dict):
        # ✅ Find the maximum list length
        max_length = max((len(v) for v in data.values() if isinstance(v, list)), default=0)

        # ✅ Normalize lists to match the maximum length
        for key, value in data.items():
            if isinstance(value, list):
                data[key] = list(itertools.islice(value + [""] * max_length, max_length))  # Fill missing slots with ""

    return data

def format_excel(ws, current_file=None):
    """Applies professional formatting to the Excel sheet."""
    header_font = Font(bold=True, name="Arial", size=12)
    body_font = Font(name="Arial", size=11)
    center_aligned_text = Alignment(horizontal="center")
    right_aligned_text = Alignment(horizontal="right")
    left_aligned_text = Alignment(horizontal="left", wrap_text=True)
    border_style = Border(left=Side(style='thin', color="D3D3D3"),
                         right=Side(style='thin', color="D3D3D3"),
                         top=Side(style='thin', color="D3D3D3"),
                         bottom=Side(style='thin', color="D3D3D3"))
    fill_color = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    # Format headers
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_aligned_text
        cell.border = border_style
        cell.fill = fill_color

    # Format body cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.border = border_style
            
            # First column (Metric names) always left aligned
            if cell.column == 1:
                cell.alignment = left_aligned_text
            else:
                # For data columns, right align if it's a number or percentage
                if cell.value and isinstance(cell.value, str):
                    if any(c.isdigit() for c in cell.value):
                        cell.alignment = right_aligned_text
                    else:
                        cell.alignment = left_aligned_text
                else:
                    cell.alignment = right_aligned_text

    # Adjust column widths
    for col in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col)
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in ws[column_letter])
        ws.column_dimensions[column_letter].width = max(18, max_length + 2)

def create_excel_file(CURRENT_PATH):
    # ✅ Define paths dynamically
    json_folder = os.path.join(CURRENT_PATH, "json_files")
    excel_folder = os.path.join(CURRENT_PATH, "excel_files")

    # ✅ Ensure output folder exists
    os.makedirs(excel_folder, exist_ok=True)

    # ✅ Generate the Excel file name dynamically
    excel_filename = get_excel_filename(json_folder)

    # ✅ Define full path for Excel file
    target_excel = os.path.join(excel_folder, excel_filename)

    # ✅ Get all JSON files in the folder
    json_files = [f for f in os.listdir(json_folder) if f.endswith(".json")]

    if not json_files:
        print(f"❌ ERROR: No JSON files found in {json_folder}")
        sys.exit(1)

    # Define the desired order of sheets
    desired_order = [
        "summary.json",
        "analysis.json",
        "income_statement.json",
        "adjustments.json",
        "balance_sheet.json",
        "company_info.json"
    ]

    # Sort the JSON files based on the desired order
    def get_file_order(file):
        try:
            return desired_order.index(file)
        except ValueError:
            # Files not in the desired order list will be placed at the end
            return len(desired_order)

    # Sort the json_files according to the desired order
    sorted_json_files = sorted(json_files, key=get_file_order)

    try:
        # ✅ Create Excel Writer
        with pd.ExcelWriter(target_excel, engine="openpyxl") as writer:
            for json_file in sorted_json_files:
                json_path = os.path.join(json_folder, json_file)
                
                # ✅ Read JSON File
                with open(json_path, "r") as f:
                    data = json.load(f)
                print(f"🔍 Processing JSON: {json_file}")

                # ✅ Skip JSON files that contain "Not Found"
                if isinstance(data, dict) and ("Not Found" in data or all(value is None for value in data.values())):
                    print(f"⚠️ Skipping {json_file}: Contains 'Not Found' or all values are null")
                    continue  

                # ✅ Skip files where all values are `None`, empty, or 'Not Available'
                if isinstance(data, dict) and all(value in [None, "", "Not Available"] for value in data.values()):
                    print(f"⚠️ Skipping {json_file}: All values are empty, None, or 'Not Available'")
                    continue  

                # ✅ Convert flat dictionary (e.g., company_info.json) into a table format
                if isinstance(data, dict) and all(not isinstance(v, (dict, list)) for v in data.values()):
                    data = [{"Field": k, "Value": v} for k, v in data.items()]  

                data = normalize_json_data(data)  # ✅ Normalize data before converting to DataFrame
                df = pd.DataFrame(data)
        

                # ✅ Skip DataFrames where all values are `None`, empty, or "Not Available"
                if df.replace("", None).isna().all().all():
                    print(f"⚠️ Skipping {json_file}: All values are empty or 'Not Available'")
                    continue  

                # ✅ Handle Specific Formatting
                if "Years" in df.columns:
                    df = df.set_index("Years").transpose()
                    df = df[sorted(df.columns)]  
                    df.insert(0, "Metric", df.index)  
                    df.reset_index(drop=True, inplace=True)

                    # Convert string percentages and numbers to float
                    for col in df.columns[1:]:
                        df[col] = df[col].apply(lambda x: 
                            float(str(x).strip('%')) if isinstance(x, str) and '%' in x 
                            else float(str(x).replace(',', '')) if isinstance(x, (str)) and ',' in str(x)
                            else float(x) if isinstance(x, (int, float, str)) and str(x).strip() not in ['', 'nan']
                            else 0
                        )

                    # Format numbers based on the metric type
                    for idx, row in df.iterrows():
                        metric = row['Metric']
                        for col in df.columns[1:]:
                            value = row[col]
                            if isinstance(value, (int, float)):
                                if 'Ratio' in metric or 'multiple' in metric:
                                    df.at[idx, col] = f"{value:.2f}"
                                elif '%' in metric:
                                    df.at[idx, col] = f"{value:.1f}%"
                                elif 'Days' in metric:
                                    df.at[idx, col] = f"{value:.1f}"
                                elif 'per employee ($)' in metric or 'Revenue' in metric or 'Income' in metric:
                                    df.at[idx, col] = f"{value:,.0f}"
                                else:
                                    df.at[idx, col] = f"{value:.2f}"

                # ✅ Define sheet name dynamically
                sheet_name = json_file.replace(".json", "").replace("_", " ").title()[:31]  

                # ✅ Write DataFrame to Excel Sheet
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                # ✅ Format Excel Sheet
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                format_excel(worksheet)  # ✅ Apply formatting function

            pass # ✅ End of Excel Writer block

       # Verify the file was created
        if os.path.exists(target_excel):
            print(f"✅ Excel file '{target_excel}' has been created successfully!")
            print("Now returning:", target_excel)
            return target_excel  # Return the path here
        else:
            print(f"❌ Failed to create Excel file at '{target_excel}'")
            return None
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
        return None
