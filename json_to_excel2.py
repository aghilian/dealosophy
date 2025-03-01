import sys
import os
import json
import pandas as pd
import openpyxl
import datetime
import subprocess
import itertools
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ✅ Get user_email and user_history_count from command-line arguments
if len(sys.argv) < 3:
    print("Error: Missing arguments. Usage: json_to_excel.py <user_email> <user_history_count>")
    sys.exit(1)

user_email = sys.argv[1]
user_history_count = str(sys.argv[2])  # Ensure it's a string for paths

# Get optional message_id and subject for threaded email replies
message_id = sys.argv[3] if len(sys.argv) > 3 else None
subject = sys.argv[4] if len(sys.argv) > 4 else None

print("✝️ Message info received from extract_data.py", user_email, user_history_count, message_id, subject)

def get_excel_filename(json_folder):
    """
    Determines the Excel file name based on company_info.json.
    If the company name is available, uses it.
    Otherwise, names the file as company_data_[Date]_[Time].xlsx.
    """
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    excel_filename = f"company_data_{timestamp}.xlsx"

    company_info_path = os.path.join(json_folder, "company_info.json")
    if os.path.exists(company_info_path):
        with open(company_info_path, "r") as f:
            company_data = json.load(f)

        if isinstance(company_data, dict) and "Name" in company_data and company_data["Name"]:
            sanitized_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in company_data["Name"])  
            excel_filename = f"{sanitized_name}.xlsx"

    print(f"📁 Excel file will be saved as: {excel_filename}")
    return excel_filename

def normalize_json_data(data):
    """
    Ensures all lists in the JSON data have the same length by filling missing values with empty strings.
    """
    if isinstance(data, dict):
        # ✅ Find the maximum list length
        max_length = max((len(v) for v in data.values() if isinstance(v, list)), default=0)

        # ✅ Normalize lists to match the maximum length
        for key, value in data.items():
            if isinstance(value, list):
                data[key] = list(itertools.islice(value + [""] * max_length, max_length))  # Fill missing slots with ""

    return data

def format_excel(ws):
    """Applies professional formatting to the Excel sheet, including wrap text for long fields."""
    header_font = Font(bold=True, name="Arial", size=12)
    body_font = Font(name="Arial", size=11)
    center_aligned_text = Alignment(horizontal="center")
    right_aligned_text = Alignment(horizontal="right")
    left_aligned_text = Alignment(horizontal="left", wrap_text=True)
    border_style = Border(left=Side(style='thin', color="D3D3D3"),
                          right=Side(style='thin', color="D3D3D3"),
                          top=Side(style='thin', color="D3D3D3"),
                          bottom=Side(style='thin', color="D3D3D3"))
    fill_color = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    # ✅ Format headers
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_aligned_text
        cell.border = border_style
        cell.fill = fill_color

    # ✅ Format body cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            if isinstance(cell.value, (int, float)):
                if cell.value == 0:
                    cell.number_format = " "  
                else:
                    cell.number_format = "#,##0;(#,##0)"
                cell.alignment = right_aligned_text
            else:
                cell.alignment = left_aligned_text

    # ✅ Apply wrap text for fields in column A if they exceed 30 characters
    for row in ws.iter_rows(min_row=2, max_col=1):
        for cell in row:
            if isinstance(cell.value, str) and len(cell.value) > 30:
                cell.alignment = Alignment(wrap_text=True)

    # ✅ Adjust column widths dynamically
    for col in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col)
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in ws[column_letter])
        ws.column_dimensions[column_letter].width = max(18, max_length + 2)

# ✅ Define paths dynamically
json_folder = os.path.join("users", user_email, user_history_count, "json_files")
excel_folder = os.path.join("users", user_email, user_history_count, "excel_files")

# ✅ Ensure output folder exists
os.makedirs(excel_folder, exist_ok=True)

# ✅ Get all JSON files in the folder
json_files = [f for f in os.listdir(json_folder) if f.endswith(".json")]

if not json_files:
    print(f"❌ ERROR: No JSON files found in {json_folder}")
    sys.exit(1)

# ✅ Generate the Excel file name dynamically
excel_filename = get_excel_filename(json_folder)

# ✅ Define full path for Excel file
target_excel = os.path.join(excel_folder, excel_filename)

# ✅ Create Excel Writer
with pd.ExcelWriter(target_excel, engine="openpyxl") as writer:
    for json_file in json_files:
        json_path = os.path.join(json_folder, json_file)
        
        # ✅ Read JSON File
        with open(json_path, "r") as f:
            data = json.load(f)

        print(f"🔍 Processing JSON: {json_file}")

        # ✅ Skip JSON files that contain "Not Found"
        if isinstance(data, dict) and "Not Found" in data:
            print(f"⚠️ Skipping {json_file}: Contains 'Not Found'")
            continue  

        # ✅ Skip files where all values are `None`, empty, or 'Not Available'
        if isinstance(data, dict) and all(value in [None, "", "Not Available"] for value in data.values()):
            print(f"⚠️ Skipping {json_file}: All values are empty, None, or 'Not Available'")
            continue  

        # ✅ Convert flat dictionary (e.g., company_info.json) into a table format
        if isinstance(data, dict) and all(not isinstance(v, (dict, list)) for v in data.values()):
            data = [{"Field": k, "Value": v} for k, v in data.items()]  

        data = normalize_json_data(data)  # ✅ Normalize data before converting to DataFrame
        df = pd.DataFrame(data)
  

        # ✅ Skip DataFrames where all values are `None`, empty, or "Not Available"
        if df.replace("", None).isna().all().all():
            print(f"⚠️ Skipping {json_file}: All values are empty or 'Not Available'")
            continue  

        # ✅ Handle Specific Formatting
        if "Years" in df.columns:
            df = df.set_index("Years").transpose()
            df = df[sorted(df.columns)]  
            df.insert(0, "Metric", df.index)  
            df.reset_index(drop=True, inplace=True)

            for col in df.columns[1:]:  
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        # ✅ Define sheet name dynamically
        sheet_name = json_file.replace(".json", "").replace("_", " ").title()[:31]  

        # ✅ Write DataFrame to Excel Sheet
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # ✅ Format Excel Sheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        format_excel(worksheet)  # ✅ Apply formatting function

print(f"✅ Excel file '{target_excel}' has been created successfully!")

# ✅ Send results via email
print("📧 Sending results via email...")
subprocess.run(["python", "send_results2.py", user_email, target_excel, message_id, subject ])

print("🚣‍♀️ Message info sent to send_results2.py", user_email, target_excel, message_id, subject)
