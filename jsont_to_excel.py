import json
import pandas as pd
import sys
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ✅ Get user_email and user_history_count from command-line arguments
if len(sys.argv) < 3:
    print("Error: Missing arguments. Usage: json_to_excel.py <user_email> <user_history_count>")
    sys.exit(1)

user_email = sys.argv[1]
user_history_count = str(sys.argv[2])  # Ensure it's a string for paths

# ✅ Define paths dynamically
json_folder = os.path.join("users", user_email, user_history_count, "json_files")
excel_folder = os.path.join("users", user_email, user_history_count, "excel_files")


# ✅ Ensure output folder exists
os.makedirs(excel_folder, exist_ok=True)

# ✅ Use dynamic path for JSON file
data_path = os.path.join(json_folder, "income_statement.json")

if not os.path.exists(data_path):
    print(f"❌ ERROR: JSON file not found: {data_path}")
    sys.exit(1)

with open(data_path, "r") as f:
    data = json.load(f)

# ✅ Define dynamic output file
target_excel = os.path.join(excel_folder, "Financial_Statement.xlsx")

def format_excel(ws):
    """Applies professional formatting to the Excel sheet."""
    header_font = Font(bold=True, name="Arial", size=12)
    body_font = Font(name="Arial", size=11)
    center_aligned_text = Alignment(horizontal="center")
    right_aligned_text = Alignment(horizontal="right")
    left_aligned_text = Alignment(horizontal="left")
    border_style = Border(left=Side(style='thin', color="D3D3D3"),
                          right=Side(style='thin', color="D3D3D3"),
                          top=Side(style='thin', color="D3D3D3"),
                          bottom=Side(style='thin', color="D3D3D3"))
    fill_color = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    
    # Format headers
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_aligned_text
        cell.border = border_style
        cell.fill = fill_color
    
    # Format body cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            if isinstance(cell.value, (int, float)):
                if cell.value == 0:
                    cell.number_format = " "  # Make zero values appear as empty
                else:
                    cell.number_format = "#,##0;(#,##0)"
                cell.alignment = right_aligned_text
            else:
                cell.alignment = left_aligned_text

 # Adjust column widths dynamically
    for col in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col)
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in ws[column_letter])
        ws.column_dimensions[column_letter].width = max(18, max_length + 2)

# Load JSON file
data_path = "income_statement.json"
with open(data_path, "r") as f:
    data = json.load(f)

# Convert JSON data to DataFrame
df = pd.DataFrame(data)
df = df.set_index("Years").transpose()  # Transpose the DataFrame so Years are columns
df = df[sorted(df.columns)]  # Ensure years are sorted in ascending order

df.insert(0, "Metric", df.index)  # Restore first column with row names

df.reset_index(drop=True, inplace=True)  # Reset index to keep it clean

# Convert all values to numeric where possible and handle non-numeric values
for col in df.columns[1:]:  # Skip the first column (Metric names)
    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)  # Convert to numeric, replacing NaN with 0
    #df[col] = df[col] / 1000  # Convert to thousands

# Insert blank rows where "Gap" appears, including "Gap1", "Gap2", etc.
def insert_blank_rows(df):
    new_df = pd.DataFrame()
    for index, row in df.iterrows():
        if row["Metric"].startswith("Gap"):
            blank_row = pd.Series(["" for _ in row], index=row.index)
            new_df = pd.concat([new_df, blank_row.to_frame().T])
        else:
            new_df = pd.concat([new_df, row.to_frame().T])
    return new_df

df = insert_blank_rows(df)

# Create an Excel writer and save DataFrame
target_excel = "Financial_Statement.xlsx"
with pd.ExcelWriter(target_excel, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Income Statement", index=False)
    workbook = writer.book
    worksheet = writer.sheets["Income Statement"]
    format_excel(worksheet)

print(f"Excel file '{target_excel}' has been created successfully!")
